import json
import math
import os
import traceback
import customtkinter as ctk
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")


ZONES = [2, 3, 4, 5, 6, 7, 8, 9, 17]
SERVICES = ["Ground", "Home Delivery"]
AHS_TYPES = ["Weight", "Package", "Dimension"]

DAS_TYPES = [
    "DAS Comm",
    "DAS Extended Comm",
    "DAS Resi",
    "DAS Extended Resi",
    "DAS Remote Commercial",
    "DAS Remote Residential",
    "Hawaii - Commercial",
    "Hawaii - Residential",
    "Alaska - Commercial",
    "Alaska - Residential",
]

SIGNATURE_TYPES = [
    "Adult Signature",
    "Direct Signature",
    "Indirect Signature",
]

DEFAULT_CONFIG_PATH = "billing_tool_config.json"
BASE_RATE_MIN_WEIGHT = 1
BASE_RATE_MAX_WEIGHT = 150


class UserInputError(Exception):
    pass


class RateTableFrame(ttk.LabelFrame):
    """Editable rate table with both vertical and horizontal scrolling."""
    def __init__(self, parent, title: str, columns: List[str], height: int = 10):
        super().__init__(parent, text=title)
        self.columns = columns

        table_shell = ttk.Frame(self)
        table_shell.pack(fill="both", expand=True, padx=6, pady=(6, 2))
        self.tree = ttk.Treeview(table_shell, columns=columns, show="headings", height=height)
        y_scroll = ttk.Scrollbar(table_shell, orient="vertical", command=self.tree.yview)
        x_scroll = ttk.Scrollbar(table_shell, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        table_shell.rowconfigure(0, weight=1)
        table_shell.columnconfigure(0, weight=1)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=98, minwidth=78, anchor="center")

        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", padx=6, pady=(3, 7))
        ttk.Button(btn_row, text="＋ Add row", command=self.add_row).pack(side="left", padx=3)
        ttk.Button(btn_row, text="Edit selected", command=self.edit_selected_row).pack(side="left", padx=3)
        ttk.Button(btn_row, text="Delete", command=self.delete_selected_row).pack(side="left", padx=3)
        ttk.Button(btn_row, text="Clear table", command=self.clear).pack(side="left", padx=3)
        ttk.Button(btn_row, text="Save", command=self.save_table).pack(side="right", padx=3)

    def add_row(self, values: List[str] = None):
        if values is None:
            values = [""] * len(self.columns)
        dialog = RowEditorDialog(self, f"Add row · {self.cget('text')}", self.columns, values)
        self.wait_window(dialog)
        if dialog.result:
            self.tree.insert("", "end", values=dialog.result)

    def edit_selected_row(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Select a row", "Please choose a row to edit first.")
            return
        item = selected[0]
        old_values = list(self.tree.item(item, "values"))
        dialog = RowEditorDialog(self, f"Edit row · {self.cget('text')}", self.columns, old_values)
        self.wait_window(dialog)
        if dialog.result:
            self.tree.item(item, values=dialog.result)

    def delete_selected_row(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Select a row", "Please choose a row to delete first.")
            return
        for item in selected:
            self.tree.delete(item)

    def save_table(self):
        messagebox.showinfo("Saved", f"{self.cget('text')} is saved in the active configuration.")

    def clear(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def load_rows(self, rows: List[List[str]]):
        self.clear()
        for row in rows:
            self.tree.insert("", "end", values=row)

    def get_rows(self) -> List[List[str]]:
        return [list(self.tree.item(item, "values")) for item in self.tree.get_children()]


class ScrollablePage(ttk.Frame):
    """A canvas-backed page wrapper so long settings pages stay usable on small screens."""
    def __init__(self, parent):
        super().__init__(parent)
        self.canvas = tk.Canvas(self, highlightthickness=0, background="#FFF8FB")
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.content = ttk.Frame(self.canvas)
        self.window = self.canvas.create_window((0, 0), window=self.content, anchor="nw")
        self.content.bind("<Configure>", self._sync_scroll_region)
        self.canvas.bind("<Configure>", self._fit_content_width)
        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)

    def _sync_scroll_region(self, _event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _fit_content_width(self, event):
        self.canvas.itemconfigure(self.window, width=event.width)

    def _bind_mousewheel(self, _event=None):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, _event=None):
        self.canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


class RowEditorDialog(tk.Toplevel):
    DROPDOWN_OPTIONS = {
        "Service": SERVICES,
        "AHS Type": AHS_TYPES,
        "DAS Type": DAS_TYPES,
        "Signature Type": SIGNATURE_TYPES,
        "Zone": [str(zone) for zone in ZONES],
    }

    def __init__(self, parent, title: str, columns: List[str], values: List[str]):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.entries = []
        self.transient(parent)
        self.grab_set()
        self.configure(bg="#FFF8FB")

        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, padx=14, pady=12)
        ttk.Label(frm, text="Use dropdowns for recognised carrier values; type custom fees or dates directly.", style="Hint.TLabel").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        for i, (col, val) in enumerate(zip(columns, values), start=1):
            ttk.Label(frm, text=col).grid(row=i, column=0, sticky="w", padx=(0, 10), pady=4)
            if col in self.DROPDOWN_OPTIONS:
                entry = ttk.Combobox(frm, values=self.DROPDOWN_OPTIONS[col], state="readonly", width=27)
                entry.set(str(val) if str(val) else self.DROPDOWN_OPTIONS[col][0])
            else:
                entry = ttk.Entry(frm, width=30)
                entry.insert(0, str(val))
            entry.grid(row=i, column=1, sticky="ew", pady=4)
            self.entries.append(entry)

        btns = ttk.Frame(frm)
        btns.grid(row=len(columns) + 1, column=0, columnspan=2, pady=(12, 0))
        ttk.Button(btns, text="Save row", style="Accent.TButton", command=self.on_save).pack(side="left", padx=4)
        ttk.Button(btns, text="Cancel", command=self.destroy).pack(side="left", padx=4)
        frm.columnconfigure(1, weight=1)

    def on_save(self):
        self.result = [entry.get().strip() for entry in self.entries]
        self.destroy()


class FedExRepricingTool:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("FedEx Repricing Studio")
        self.root.geometry("1240x780")
        self.root.minsize(1060, 660)
        self.root.configure(fg_color="#f5f3f7")
        # Apply a FedEx deep-purple Windows title bar after the window is created.
        self.root.after(80, self._set_windows_title_bar_color)

        self.billing_path = tk.StringVar()
        self.output_path = tk.StringVar(value=os.path.join(os.getcwd(), "repriced_result.xlsx"))
        self.config_path = tk.StringVar(value=os.path.join(os.getcwd(), DEFAULT_CONFIG_PATH))

        self.dim_factor = tk.DoubleVar(value=139.0)
        self.fuel_percent = tk.DoubleVar(value=18.75)

        self.use_oversize_longest = tk.BooleanVar(value=False)
        self.use_oversize_weight = tk.BooleanVar(value=True)
        self.use_oversize_cubic = tk.BooleanVar(value=False)
        self.use_oversize_lg = tk.BooleanVar(value=True)

        self.oversize_longest_side = tk.DoubleVar(value=96.0)
        self.oversize_actual_weight = tk.DoubleVar(value=110.0)
        self.oversize_cubic_inches = tk.DoubleVar(value=17280.0)
        self.length_girth_limit = tk.DoubleVar(value=130.0)
        self.oversize_min_billable_weight = tk.DoubleVar(value=90.0)

        self.ahs_weight_threshold = tk.DoubleVar(value=50.0)
        self.ahs_dimension_lg_limit = tk.DoubleVar(value=105.0)
        self.ahs_dimension_longest_side = tk.DoubleVar(value=48.0)
        self.ahs_dimension_second_side = tk.DoubleVar(value=30.0)
        self.ahs_dimension_cubic_inches = tk.DoubleVar(value=10368.0)
        self.ahs_min_billable_weight = tk.DoubleVar(value=40.0)

        self._build_ui()
        self._load_default_tables()

    def _set_windows_title_bar_color(self):
        """Use a deep-purple native title bar on supported Windows versions."""
        if os.name != "nt":
            return

        try:
            import ctypes

            self.root.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(self.root.winfo_id())
            if not hwnd:
                hwnd = self.root.winfo_id()

            # Windows COLORREF uses BGR order.
            caption_color = ctypes.c_int(0x004F0B2E)  # #2E0B4F
            border_color = ctypes.c_int(0x004F0B2E)   # #2E0B4F
            text_color = ctypes.c_int(0x00FFFFFF)     # white

            dwm = ctypes.windll.dwmapi.DwmSetWindowAttribute
            dwm(hwnd, 35, ctypes.byref(caption_color), ctypes.sizeof(caption_color))
            dwm(hwnd, 34, ctypes.byref(border_color), ctypes.sizeof(border_color))
            dwm(hwnd, 36, ctypes.byref(text_color), ctypes.sizeof(text_color))
        except Exception:
            # Older Windows versions may not support custom title-bar colors.
            pass

    def _build_ui(self):
        """Build a workflow-first desktop UI without changing repricing logic."""
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        self.colors = {
            "purple": "#7A4E85",
            "deep": "#493151",
            "orange": "#F1846B",
            "ink": "#3B2D41",
            "muted": "#8B778F",
            "canvas": "#FFF8FB",
            "card": "#FFFFFF",
            "line": "#E9D9E8",
            "soft": "#F9EAF5",
        }
        self.root.title("FedEx Repricing Studio")
        self.root.geometry("1240x780")
        self.root.minsize(1060, 660)
        self.root.configure(fg_color=self.colors["canvas"])

        style.configure("TFrame", background=self.colors["canvas"])
        style.configure("TLabel", background=self.colors["canvas"], foreground=self.colors["ink"], font=("Segoe UI", 10))
        style.configure("PageTitle.TLabel", background=self.colors["canvas"], foreground=self.colors["purple"], font=("Segoe UI", 16, "bold"))
        style.configure("Hint.TLabel", background=self.colors["canvas"], foreground=self.colors["muted"], font=("Segoe UI", 9))
        style.configure("TLabelframe", background=self.colors["card"], bordercolor=self.colors["line"], relief="solid", borderwidth=1)
        style.configure("TLabelframe.Label", background=self.colors["card"], foreground=self.colors["purple"], font=("Segoe UI", 10, "bold"))
        style.configure("TEntry", fieldbackground="#FFFFFF", foreground=self.colors["ink"], bordercolor=self.colors["line"], relief="solid", borderwidth=1)
        style.configure("TCombobox", fieldbackground="#FFFFFF", background="#FFFFFF", foreground=self.colors["ink"], bordercolor=self.colors["line"], arrowcolor=self.colors["purple"])
        style.configure("TButton", padding=[8, 5], font=("Segoe UI", 9, "bold"), foreground=self.colors["purple"])
        style.configure("Accent.TButton", padding=[10, 6], font=("Segoe UI", 9, "bold"), foreground="#FFFFFF", background=self.colors["orange"])
        style.map("Accent.TButton", background=[("active", "#D95700")])
        style.configure("Treeview", background="#FFFFFF", foreground=self.colors["ink"], rowheight=24, fieldbackground="#FFFFFF", borderwidth=1, bordercolor=self.colors["line"], font=("Segoe UI", 9))
        style.configure("Treeview.Heading", background=self.colors["purple"], foreground="#FFFFFF", font=("Segoe UI", 9, "bold"), relief="flat")
        style.configure("TNotebook", background=self.colors["canvas"], borderwidth=0)
        style.configure("TNotebook.Tab", background="#EEE8F2", foreground=self.colors["purple"], padding=[12, 7], font=("Segoe UI", 9, "bold"))
        style.map("TNotebook.Tab", background=[("selected", self.colors["purple"])], foreground=[("selected", "white")])

        shell = ctk.CTkFrame(self.root, fg_color=self.colors["canvas"], corner_radius=0)
        shell.pack(fill="both", expand=True)

        sidebar = ctk.CTkFrame(shell, width=236, fg_color=self.colors["deep"], corner_radius=0)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        ctk.CTkLabel(sidebar, text="FEDEX", font=("Segoe UI", 19, "bold"), text_color="#FFFFFF").pack(anchor="w", padx=24, pady=(28, 0))
        ctk.CTkLabel(sidebar, text="REPRICING STUDIO  ♡", font=("Segoe UI", 10, "bold"), text_color="#D8C6F0").pack(anchor="w", padx=24, pady=(0, 24))
        ctk.CTkLabel(sidebar, text="WORKFLOW  ✦", font=("Segoe UI", 9, "bold"), text_color="#BCA8D2").pack(anchor="w", padx=24, pady=(0, 8))

        self.nav_buttons = {}
        self.pages = {}
        self.page_titles = {
            "intake": ("1. Import & setup", "Choose billing data and keep a reusable rate configuration."),
            "rules": ("2. Rating rules", "Manage DIM, fuel, oversize and AHS thresholds in one place."),
            "base": ("3. Base rates", "Review a service rate grid and use dedicated templates."),
            "surcharges": ("4. Surcharges", "Maintain accessorial, residential, signature and demand charges."),
            "run": ("5. Reprice & review", "Run the engine, scan the output and export the workbook."),
        }
        nav_items = [
            ("intake", "01  ✦ Import & setup"),
            ("rules", "02  ♡ Rating rules"),
            ("base", "03  ✧ Base rates"),
            ("surcharges", "04  ♡ Surcharges"),
            ("run", "05  ✦ Reprice & review"),
        ]
        for name, label in nav_items:
            button = ctk.CTkButton(
                sidebar, text=label, anchor="w", height=40, corner_radius=8,
                font=("Segoe UI", 11, "bold"), fg_color="transparent", hover_color="#785380",
                text_color="#F6F0FF", command=lambda key=name: self._show_page(key)
            )
            button.pack(fill="x", padx=14, pady=3)
            self.nav_buttons[name] = button

        sidebar_footer = ctk.CTkFrame(sidebar, fg_color="#604568", corner_radius=10)
        sidebar_footer.pack(side="bottom", fill="x", padx=14, pady=18)
        ctk.CTkLabel(sidebar_footer, text="TIP", font=("Segoe UI", 9, "bold"), text_color="#D8C6F0").pack(anchor="w", padx=12, pady=(10, 0))
        ctk.CTkLabel(sidebar_footer, text="Save a configuration after updating any rate table.", wraplength=180, justify="left", font=("Segoe UI", 10), text_color="#FFFFFF").pack(anchor="w", padx=12, pady=(3, 11))

        body = ctk.CTkFrame(shell, fg_color=self.colors["canvas"], corner_radius=0)
        body.pack(side="left", fill="both", expand=True)
        header = ctk.CTkFrame(body, fg_color="transparent", corner_radius=0)
        header.pack(fill="x", padx=30, pady=(24, 8))
        self.page_title_label = ctk.CTkLabel(header, text="", font=("Segoe UI", 22, "bold"), text_color=self.colors["purple"])
        self.page_title_label.pack(anchor="w")
        self.page_subtitle_label = ctk.CTkLabel(header, text="", font=("Segoe UI", 11), text_color=self.colors["muted"])
        self.page_subtitle_label.pack(anchor="w", pady=(2, 0))

        self.page_host = ctk.CTkFrame(body, fg_color="transparent", corner_radius=0)
        self.page_host.pack(fill="both", expand=True, padx=24, pady=(0, 20))

        self.intake_scroll = ScrollablePage(self.page_host)
        self.rules_scroll = ScrollablePage(self.page_host)
        self.base_scroll = ScrollablePage(self.page_host)
        self.surcharges_scroll = ScrollablePage(self.page_host)
        self.run_scroll = ScrollablePage(self.page_host)
        self.tab_files = self.intake_scroll.content
        self.tab_rules = self.rules_scroll.content
        self.tab_base_rates = self.base_scroll.content
        self.tab_surcharges = self.surcharges_scroll.content
        self.tab_run = self.run_scroll.content
        self.pages = {
            "intake": self.intake_scroll,
            "rules": self.rules_scroll,
            "base": self.base_scroll,
            "surcharges": self.surcharges_scroll,
            "run": self.run_scroll,
        }

        self._build_files_tab()
        self._build_rules_tab()
        self._build_base_rates_tab()
        self._build_surcharges_tab()
        self._build_run_tab()
        self._show_page("intake")

    def _show_page(self, name: str):
        for page in self.pages.values():
            page.pack_forget()
        self.pages[name].pack(fill="both", expand=True)
        title, subtitle = self.page_titles[name]
        self.page_title_label.configure(text=title)
        self.page_subtitle_label.configure(text=subtitle)
        for key, button in self.nav_buttons.items():
            button.configure(fg_color="#6B2BA1" if key == name else "transparent")

    def _build_files_tab(self):
        top = ttk.LabelFrame(self.tab_files, text="Billing file")
        top.pack(fill="x", padx=2, pady=(2, 10))
        ttk.Label(top, text="Source invoice / billing data", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))
        ttk.Entry(top, textvariable=self.billing_path, width=82).grid(row=1, column=0, sticky="ew", padx=(12, 8), pady=(0, 8))
        ttk.Button(top, text="Browse billing file", command=self._pick_billing).grid(row=1, column=1, sticky="e", padx=(0, 12), pady=(0, 8))
        top.columnconfigure(0, weight=1)
        self.billing_file_note = tk.StringVar(value="No source file selected yet. Supported formats: .xlsx, .xls and .csv.")
        ttk.Label(top, textvariable=self.billing_file_note, style="Hint.TLabel").grid(row=2, column=0, columnspan=2, sticky="w", padx=12, pady=(0, 12))

        config = ttk.LabelFrame(self.tab_files, text="Reusable configuration")
        config.pack(fill="x", padx=2, pady=10)
        ttk.Label(config, text="Rate configuration", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))
        ttk.Entry(config, textvariable=self.config_path, width=82).grid(row=1, column=0, sticky="ew", padx=(12, 8), pady=(0, 8))
        ttk.Button(config, text="Choose location", command=self._pick_config).grid(row=1, column=1, padx=4, pady=(0, 8))
        ttk.Button(config, text="Load config", command=self.load_config).grid(row=1, column=2, padx=4, pady=(0, 8))
        ttk.Button(config, text="Save config", style="Accent.TButton", command=self.save_config).grid(row=1, column=3, padx=(4, 12), pady=(0, 8))
        ttk.Label(config, text="Configurations preserve all rules and rate tables, so you can reuse a validated setup for later invoices.", style="Hint.TLabel").grid(row=2, column=0, columnspan=4, sticky="w", padx=12, pady=(0, 12))
        config.columnconfigure(0, weight=1)

        templates = ttk.LabelFrame(self.tab_files, text="Start here")
        templates.pack(fill="x", padx=2, pady=10)
        ttk.Label(templates, text="New to the tool?", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 2))
        ttk.Label(templates, text="Download the sample to confirm required columns before importing live billing data.", style="Hint.TLabel").grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))
        ttk.Button(templates, text="Download billing sample", command=self.export_billing_sample).grid(row=0, column=1, rowspan=2, padx=12, pady=12)
        templates.columnconfigure(0, weight=1)

    def _build_rules_tab(self):
        container = ttk.Frame(self.tab_rules)
        container.pack(fill="both", expand=True, padx=2, pady=2)

        global_frame = ttk.LabelFrame(container, text="Core calculations")
        global_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(global_frame, text="DIM factor").grid(row=0, column=0, sticky="w", padx=(12, 6), pady=12)
        ttk.Combobox(global_frame, textvariable=self.dim_factor, values=("139", "166"), width=10).grid(row=0, column=1, sticky="w", padx=(0, 24), pady=12)
        ttk.Label(global_frame, text="Fuel surcharge (%)").grid(row=0, column=2, sticky="w", padx=(0, 6), pady=12)
        ttk.Combobox(global_frame, textvariable=self.fuel_percent, values=("15.00", "18.75", "20.00", "22.00"), width=10).grid(row=0, column=3, sticky="w", padx=(0, 12), pady=12)

        oversize_frame = ttk.LabelFrame(container, text="Oversize thresholds")
        oversize_frame.pack(fill="x", pady=10)
        oversize_fields = [
            ("Longest side", self.use_oversize_longest, self.oversize_longest_side),
            ("Actual weight", self.use_oversize_weight, self.oversize_actual_weight),
            ("Cubic inches", self.use_oversize_cubic, self.oversize_cubic_inches),
            ("Length + girth", self.use_oversize_lg, self.length_girth_limit),
        ]
        for row, (label, enabled, value) in enumerate(oversize_fields):
            ttk.Checkbutton(oversize_frame, text=f"Enable {label}", variable=enabled).grid(row=row, column=0, sticky="w", padx=12, pady=5)
            ttk.Entry(oversize_frame, textvariable=value, width=12).grid(row=row, column=1, sticky="w", padx=10, pady=5)
        ttk.Label(oversize_frame, text="Minimum billable weight").grid(row=4, column=0, sticky="w", padx=12, pady=(5, 12))
        ttk.Entry(oversize_frame, textvariable=self.oversize_min_billable_weight, width=12).grid(row=4, column=1, sticky="w", padx=10, pady=(5, 12))

        ahs_frame = ttk.LabelFrame(container, text="Additional Handling Surcharge thresholds")
        ahs_frame.pack(fill="x", pady=10)
        ahs_fields = [
            ("Weight threshold", self.ahs_weight_threshold),
            ("Length + girth limit", self.ahs_dimension_lg_limit),
            ("Longest side", self.ahs_dimension_longest_side),
            ("Second-longest side", self.ahs_dimension_second_side),
            ("Cubic inches", self.ahs_dimension_cubic_inches),
            ("Minimum billable weight", self.ahs_min_billable_weight),
        ]
        for index, (label, value) in enumerate(ahs_fields):
            row, column = divmod(index, 2)
            base = column * 2
            ttk.Label(ahs_frame, text=label).grid(row=row, column=base, sticky="w", padx=(12 if column == 0 else 28, 6), pady=7)
            ttk.Entry(ahs_frame, textvariable=value, width=12).grid(row=row, column=base + 1, sticky="w", padx=(0, 12), pady=7)

    def _build_base_rates_tab(self):
        control = ttk.LabelFrame(self.tab_base_rates, text="Base rate workspace")
        control.pack(fill="x", padx=2, pady=(2, 10))
        self.base_rate_service = tk.StringVar(value="Ground")
        ttk.Label(control, text="Service", font=("Segoe UI", 10, "bold")).pack(side="left", padx=(12, 6), pady=10)
        service_combo = ttk.Combobox(control, textvariable=self.base_rate_service, values=SERVICES, state="readonly", width=18)
        service_combo.pack(side="left", padx=(0, 12), pady=10)
        self.current_service_label = ttk.Label(control, text="Editing: Ground", font=("Segoe UI", 10, "bold"), foreground=self.colors["purple"])
        self.current_service_label.pack(side="left", padx=(0, 20), pady=10)
        service_combo.bind("<<ComboboxSelected>>", self.on_service_changed)
        ttk.Button(control, text="Import template", command=self.import_base_rate_template).pack(side="right", padx=5, pady=10)
        ttk.Button(control, text="Export template", command=self.export_base_rate_template).pack(side="right", padx=5, pady=10)
        self.base_rate_table = RateTableFrame(self.tab_base_rates, "Rate grid · all configured services", ["Service", "Weight"] + [f"Zone {z}" for z in ZONES], height=17)
        self.base_rate_table.pack(fill="both", expand=True, padx=2, pady=2)

    def _build_surcharges_tab(self):
        intro = ttk.LabelFrame(self.tab_surcharges, text="Accessorial charge library")
        intro.pack(fill="x", padx=2, pady=(2, 8))
        ttk.Label(intro, text="Use the category tabs to keep each surcharge table focused. Import or export templates without leaving the category you are working on.", style="Hint.TLabel").pack(anchor="w", padx=12, pady=10)

        book = ttk.Notebook(self.tab_surcharges)
        book.pack(fill="both", expand=True, padx=2, pady=2)
        tab_ahs = ttk.Frame(book)
        tab_das = ttk.Frame(book)
        tab_oversize = ttk.Frame(book)
        tab_residential = ttk.Frame(book)
        tab_signature = ttk.Frame(book)
        tab_demand = ttk.Frame(book)
        book.add(tab_ahs, text="AHS")
        book.add(tab_das, text="DAS")
        book.add(tab_oversize, text="Oversize")
        book.add(tab_residential, text="Residential")
        book.add(tab_signature, text="Signature")
        book.add(tab_demand, text="Demand periods")

        self._add_template_toolbar(tab_ahs, self.export_ahs_rate_template, self.import_ahs_rate_template)
        self.ahs_rate_table = RateTableFrame(tab_ahs, "Additional Handling Surcharge", ["Service", "AHS Type", "Zone", "Fee"], height=15)
        self.ahs_rate_table.pack(fill="both", expand=True, padx=2, pady=2)

        self._add_template_toolbar(tab_das, self.export_das_rate_template, self.import_das_rate_template)
        self.das_rate_table = RateTableFrame(tab_das, "Delivery Area Surcharge", ["Service", "DAS Type", "Zone", "Fee"], height=15)
        self.das_rate_table.pack(fill="both", expand=True, padx=2, pady=2)

        self._add_template_toolbar(tab_oversize, self.export_oversize_rate_template, self.import_oversize_rate_template)
        self.oversize_rate_table = RateTableFrame(tab_oversize, "Oversize fees", ["Service", "Zone", "Fee"], height=15)
        self.oversize_rate_table.pack(fill="both", expand=True, padx=2, pady=2)

        self._add_template_toolbar(tab_residential, self.export_residential_rate_template, self.import_residential_rate_template)
        self.residential_rate_table = RateTableFrame(tab_residential, "Residential fees", ["Service", "Fee"], height=15)
        self.residential_rate_table.pack(fill="both", expand=True, padx=2, pady=2)

        self._add_template_toolbar(tab_signature, self.export_signature_rate_template, self.import_signature_rate_template)
        self.signature_rate_table = RateTableFrame(tab_signature, "Signature fees", ["Signature Type", "Fee"], height=15)
        self.signature_rate_table.pack(fill="both", expand=True, padx=2, pady=2)

        demand_book = ttk.Notebook(tab_demand)
        demand_book.pack(fill="both", expand=True, padx=2, pady=2)
        general_page = ttk.Frame(demand_book)
        ahs_page = ttk.Frame(demand_book)
        oversize_page = ttk.Frame(demand_book)
        demand_book.add(general_page, text="General demand")
        demand_book.add(ahs_page, text="AHS demand")
        demand_book.add(oversize_page, text="Oversize demand")
        self._add_template_toolbar(general_page, self.export_general_demand_template, self.import_general_demand_template)
        self.general_demand_table = RateTableFrame(general_page, "General demand fees", ["Service", "Start Date", "End Date", "Fee"], height=13)
        self.general_demand_table.pack(fill="both", expand=True, padx=2, pady=2)
        self._add_template_toolbar(ahs_page, self.export_ahs_demand_template, self.import_ahs_demand_template)
        self.ahs_demand_table = RateTableFrame(ahs_page, "AHS demand fees", ["Service", "Start Date", "End Date", "Fee"], height=13)
        self.ahs_demand_table.pack(fill="both", expand=True, padx=2, pady=2)
        self._add_template_toolbar(oversize_page, self.export_oversize_demand_template, self.import_oversize_demand_template)
        self.oversize_demand_table = RateTableFrame(oversize_page, "Oversize demand fees", ["Service", "Start Date", "End Date", "Fee"], height=13)
        self.oversize_demand_table.pack(fill="both", expand=True, padx=2, pady=2)

    def _add_template_toolbar(self, parent, export_cmd, import_cmd):
        bar = ttk.Frame(parent)
        bar.pack(fill="x", pady=(2, 6))
        ttk.Label(bar, text="Template controls", font=("Segoe UI", 9, "bold")).pack(side="left", padx=4)
        ttk.Button(bar, text="Import template", command=import_cmd).pack(side="right", padx=4)
        ttk.Button(bar, text="Export template", command=export_cmd).pack(side="right", padx=4)

    def on_service_changed(self, event=None):
        self.current_service_label.config(text=f"Editing: {self.base_rate_service.get()}")

    def _build_run_tab(self):
        top = ttk.LabelFrame(self.tab_run, text="Run status")
        top.pack(fill="x", padx=2, pady=(2, 10))
        action = ctk.CTkButton(top, text="Run repricing and export workbook", fg_color=self.colors["orange"], hover_color="#D95700", corner_radius=8, height=38, font=("Segoe UI", 11, "bold"), command=self.run_repricing)
        action.pack(side="left", padx=12, pady=12)
        self.status_var = tk.StringVar(value="Ready — select a billing file, review rates, then run.")
        self.run_summary_var = tk.StringVar(value="Preview will populate after a successful run.")
        text_area = ttk.Frame(top)
        text_area.pack(side="left", fill="x", expand=True, padx=(4, 12), pady=8)
        ttk.Label(text_area, textvariable=self.status_var, font=("Segoe UI", 10, "bold"), foreground=self.colors["purple"]).pack(anchor="w")
        ttk.Label(text_area, textvariable=self.run_summary_var, style="Hint.TLabel").pack(anchor="w", pady=(2, 0))

        results = ttk.LabelFrame(self.tab_run, text="Repricing preview")
        results.pack(fill="both", expand=True, padx=2, pady=2)
        cols = ["Service", "Zone", "Actual", "DIM", "Billable", "Base", "AHS", "DAS", "Oversize", "Residential", "Signature", "Demand", "Fuel", "Total"]
        self.preview = ttk.Treeview(results, columns=cols, show="headings", height=19)
        for column in cols:
            self.preview.heading(column, text=column)
            self.preview.column(column, width=88, minwidth=64, anchor="center")
        scrollbar = ttk.Scrollbar(results, orient="horizontal", command=self.preview.xview)
        y_scrollbar = ttk.Scrollbar(results, orient="vertical", command=self.preview.yview)
        self.preview.configure(xscrollcommand=scrollbar.set, yscrollcommand=y_scrollbar.set)
        y_scrollbar.pack(side="right", fill="y", padx=(0, 6), pady=(6, 0))
        self.preview.pack(fill="both", expand=True, padx=6, pady=(6, 0))
        scrollbar.pack(fill="x", padx=6, pady=(0, 6))

    def _pick_billing(self):
        path = filedialog.askopenfilename(filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv")])
        if path:
            self.billing_path.set(path)
            try:
                filename = os.path.basename(path)
                self.billing_file_note.set(f"Selected: {filename} · ready for review and repricing.")
            except Exception:
                pass

    def _pick_config(self):
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON", "*.json")])
        if path:
            self.config_path.set(path)

    def _load_default_tables(self):
        base_rows = []
        for service in SERVICES:
            for weight in range(BASE_RATE_MIN_WEIGHT, BASE_RATE_MAX_WEIGHT + 1):
                base_rows.append([service, str(weight)] + ["0" for _ in ZONES])
        self.base_rate_table.load_rows(base_rows)

        ahs_rows = []

        for service in SERVICES:
            for ahs_type in AHS_TYPES:
                for zone in ZONES:

                    ahs_rows.append(
                        [
                            service,
                            ahs_type,
                            zone,
                            "0"
                        ]
                    )

        self.ahs_rate_table.load_rows(ahs_rows)

        das_rows = []
        for service in SERVICES:
            for das_type in DAS_TYPES:
                for zone in ZONES:
                    das_rows.append([service, das_type, str(zone), "0"])
        self.das_rate_table.load_rows(das_rows)

        oversize_rows = []
        for service in SERVICES:
            for zone in ZONES:
                oversize_rows.append([service, str(zone), "0"])
        self.oversize_rate_table.load_rows(oversize_rows)

        self.residential_rate_table.load_rows([
            ["Ground", "2.58"],
            ["Home Delivery", "1.94"],
        ])

        self.signature_rate_table.load_rows([
            ["Adult Signature", "10.00"],
            ["Direct Signature", "5.70"],
            ["Indirect Signature", "5.70"],
        ])

        self.general_demand_table.load_rows([])
        self.ahs_demand_table.load_rows([])
        self.oversize_demand_table.load_rows([])

    @staticmethod
    def _normalize_numeric_cell(val, default=0.0):
        if pd.isna(val) or str(val).strip() == "":
            return default
        return float(str(val).replace(",", "").replace("$", "").strip())

    @staticmethod
    def _to_float(x, default=0.0):
        try:
            if x is None:
                return default
            if pd.isna(x):
                return default
            return float(str(x).replace(",", "").replace("$", "").strip())
        except Exception:
            return default

    @staticmethod
    def _normalize_service(s: str) -> str:
        value = str(s).strip().lower()
        if "home" in value:
            return "Home Delivery"
        return "Ground"

    @staticmethod
    def _flag(x) -> bool:
        if pd.isna(x):
            return False
        s = str(x).strip().lower()
        return s not in ["", "0", "n", "no", "false", "none"]

    def _raise_template_error(self, template_name: str, msg: str):
        raise UserInputError(f"{template_name} template error\n\n{msg}")

    def _export_table(self, table: RateTableFrame, title: str):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        pd.DataFrame(table.get_rows(), columns=table.columns).to_excel(path, index=False)
        messagebox.showinfo("Done", f"{title} exported:\n{path}")

    def _import_table_with_validation(self, table: RateTableFrame, title: str, expected_columns: List[str], validator):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if not path:
            return

        try:
            df = pd.read_excel(path)
            df.columns = [str(c).strip() for c in df.columns]

            if list(df.columns) != expected_columns:
                self._raise_template_error(
                    title,
                    f"Expected columns:\n{expected_columns}\n\nGot:\n{list(df.columns)}"
                )

            rows = validator(df)
            table.load_rows(rows)
            messagebox.showinfo("Done", f"{title} imported:\n{path}")

        except UserInputError:
            raise
        except Exception as e:
            self._raise_template_error(title, str(e))

    def _validate_demand_template(self, title: str, df: pd.DataFrame):
        rows = []
        for i, row in df.iterrows():
            excel_row = i + 2
            service = str(row["Service"]).strip()
            start_date = str(row["Start Date"]).strip()
            end_date = str(row["End Date"]).strip()

            if service not in SERVICES:
                self._raise_template_error(title, f"Row {excel_row}: Service must be one of {SERVICES}")

            if not self._parse_date(start_date):
                self._raise_template_error(title, f"Row {excel_row}: Start Date invalid: {start_date}")

            if not self._parse_date(end_date):
                self._raise_template_error(title, f"Row {excel_row}: End Date invalid: {end_date}")

            try:
                fee = self._normalize_numeric_cell(row["Fee"], 0.0)
            except Exception:
                self._raise_template_error(title, f"Row {excel_row}: Fee must be numeric")

            rows.append([service, start_date, end_date, str(fee)])

        return rows

    def export_base_rate_template(self):

        selected_service = self.base_rate_service.get()

        filtered_rows = [
            row
            for row in self.base_rate_table.get_rows()
            if row[0] == selected_service
        ]

        filename = f"{selected_service}_Base_Rate_template.xlsx"

        path = filedialog.asksaveasfilename(
            initialfile=filename,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        pd.DataFrame(
            filtered_rows,
            columns=self.base_rate_table.columns
        ).to_excel(path, index=False)

        messagebox.showinfo(
            "Done",
            f"{selected_service} Base Rate template exported:\n{path}"
        )

    def import_base_rate_template(self):
        def validator(df: pd.DataFrame):
            rows = []
            seen = set()
            for i, row in df.iterrows():
                excel_row = i + 2
                service = str(row["Service"]).strip()
                if service not in SERVICES:
                    self._raise_template_error("Base Rate", f"Row {excel_row}: Service must be one of {SERVICES}")

                try:
                    weight = int(float(row["Weight"]))
                except Exception:
                    self._raise_template_error("Base Rate", f"Row {excel_row}: Weight must be numeric")

                if weight < BASE_RATE_MIN_WEIGHT or weight > BASE_RATE_MAX_WEIGHT:
                    self._raise_template_error("Base Rate", f"Row {excel_row}: Weight must be between {BASE_RATE_MIN_WEIGHT} and {BASE_RATE_MAX_WEIGHT}")

                key = (service, weight)
                if key in seen:
                    self._raise_template_error("Base Rate", f"Row {excel_row}: Duplicate Service + Weight: {service}, {weight}")
                seen.add(key)

                new_row = [service, str(weight)]
                for z in ZONES:
                    col = f"Zone {z}"
                    try:
                        fee = self._normalize_numeric_cell(row[col], 0.0)
                    except Exception:
                        self._raise_template_error("Base Rate", f"Row {excel_row}: {col} must be numeric")
                    new_row.append(str(fee))
                rows.append(new_row)
            if not rows:
                self._raise_template_error("Base Rate", "Template is empty.")
            return rows

        path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xls")]
        )

        if not path:
            return

        try:
            df = pd.read_excel(path)
            df.columns = [str(c).strip() for c in df.columns]

            expected_columns = ["Service", "Weight"] + [f"Zone {z}" for z in ZONES]

            if list(df.columns) != expected_columns:
                self._raise_template_error(
                    "Base Rate",
                    f"Expected columns:\n{expected_columns}\n\nGot:\n{list(df.columns)}"
                )

            rows = validator(df)

            selected_services = sorted(set(str(r[0]).strip() for r in rows))

            existing_rows = self.base_rate_table.get_rows()

            filtered_existing = [
                r for r in existing_rows
                if str(r[0]).strip() not in selected_services
            ]

            final_rows = filtered_existing + rows

            self.base_rate_table.load_rows(final_rows)

            messagebox.showinfo(
                "Done",
                f"Base Rate imported for {', '.join(selected_services)}:\n{path}"
            )

        except UserInputError:
            raise
        except Exception as e:
            self._raise_template_error("Base Rate", str(e))


    def export_accessorial_sheet(self):

        rows = []

        # AHS
        for row in self.ahs_rate_table.get_rows():
            rows.append([
                "AHS",
                row[0],
                row[1],
                row[2],
                "",
                "",
                row[3]
            ])

        # DAS
        for row in self.das_rate_table.get_rows():
            rows.append([
                "DAS",
                row[0],
                row[1],
                row[2],
                "",
                "",
                row[3]
            ])

        # Oversize
        for row in self.oversize_rate_table.get_rows():
            rows.append([
                "Oversize",
                row[0],
                "Oversize",
                row[1],
                "",
                "",
                row[2]
            ])

        # Residential
        for row in self.residential_rate_table.get_rows():
            rows.append([
                "Residential",
                row[0],
                "Residential",
                "",
                "",
                "",
                row[1]
            ])

        # Signature
        for row in self.signature_rate_table.get_rows():
            rows.append([
                "Signature",
                "",
                row[0],
                "",
                "",
                "",
                row[1]
            ])

        # General Demand
        for row in self.general_demand_table.get_rows():
            rows.append([
                "General Demand",
                row[0],
                "Demand",
                "",
                row[1],
                row[2],
                row[3]
            ])

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        pd.DataFrame(
            rows,
            columns=[
                "Charge Category",
                "Service",
                "Sub Type",
                "Zone",
                "Start Date",
                "End Date",
                "Fee"
            ]
        ).to_excel(path, index=False)

        messagebox.showinfo(
            "Done",
            f"Accessorial Sheet exported:\n{path}"
        )

    def export_ahs_rate_template(self):

        path = filedialog.asksaveasfilename(
            initialfile="AHS_template.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        pd.DataFrame(
            self.ahs_rate_table.get_rows(),
            columns=self.ahs_rate_table.columns
        ).to_excel(path, index=False)

        messagebox.showinfo(
            "Done",
            f"AHS template exported:\n{path}"
        )

    def import_ahs_rate_template(self):
        def validator(df: pd.DataFrame):
            rows = []
            seen = set()
            for i, row in df.iterrows():
                excel_row = i + 2
                service = str(row["Service"]).strip()
                ahs_type = str(row["AHS Type"]).strip()

                if service not in SERVICES:
                    self._raise_template_error("AHS", f"Row {excel_row}: Service must be one of {SERVICES}")
                if ahs_type not in AHS_TYPES:
                    self._raise_template_error("AHS", f"Row {excel_row}: AHS Type must be one of {AHS_TYPES}")

                try:
                    zone = int(float(row["Zone"]))
                except Exception:
                    self._raise_template_error("AHS", f"Row {excel_row}: Zone must be numeric")
                if zone not in ZONES:
                    self._raise_template_error("AHS", f"Row {excel_row}: Zone must be one of {ZONES}")

                try:
                    fee = self._normalize_numeric_cell(row["Fee"], 0.0)
                except Exception:
                    self._raise_template_error("AHS", f"Row {excel_row}: Fee must be numeric")

                key = (service, ahs_type, zone)
                if key in seen:
                    self._raise_template_error("AHS", f"Row {excel_row}: Duplicate Service + AHS Type + Zone")
                seen.add(key)

                rows.append([service, ahs_type, str(zone), str(fee)])

            if not rows:
                self._raise_template_error("AHS", "Template is empty.")
            return rows

        self._import_table_with_validation(
            self.ahs_rate_table,
            "AHS",
            ["Service", "AHS Type", "Zone", "Fee"],
            validator,
        )

    def export_das_rate_template(self):

        path = filedialog.asksaveasfilename(
            initialfile="DAS_template.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        pd.DataFrame(
            self.das_rate_table.get_rows(),
            columns=self.das_rate_table.columns
        ).to_excel(path, index=False)

        messagebox.showinfo(
            "Done",
            f"DAS template exported:\n{path}"
        )

    def import_das_rate_template(self):
        def validator(df: pd.DataFrame):
            rows = []
            seen = set()
            for i, row in df.iterrows():
                excel_row = i + 2
                service = str(row["Service"]).strip()
                das_type = str(row["DAS Type"]).strip()

                if service not in SERVICES:
                    self._raise_template_error("DAS", f"Row {excel_row}: Service must be one of {SERVICES}")
                if das_type not in DAS_TYPES:
                    self._raise_template_error("DAS", f"Row {excel_row}: DAS Type must be one of {DAS_TYPES}")

                try:
                    zone = int(float(row["Zone"]))
                except Exception:
                    self._raise_template_error("DAS", f"Row {excel_row}: Zone must be numeric")
                if zone not in ZONES:
                    self._raise_template_error("DAS", f"Row {excel_row}: Zone must be one of {ZONES}")

                try:
                    fee = self._normalize_numeric_cell(row["Fee"], 0.0)
                except Exception:
                    self._raise_template_error("DAS", f"Row {excel_row}: Fee must be numeric")

                key = (service, das_type, zone)
                if key in seen:
                    self._raise_template_error("DAS", f"Row {excel_row}: Duplicate Service + DAS Type + Zone")
                seen.add(key)

                rows.append([service, das_type, str(zone), str(fee)])

            if not rows:
                self._raise_template_error("DAS", "Template is empty.")
            return rows

        self._import_table_with_validation(
            self.das_rate_table,
            "DAS",
            ["Service", "DAS Type", "Zone", "Fee"],
            validator,
        )

    def export_oversize_rate_template(self):

        path = filedialog.asksaveasfilename(
            initialfile="Oversize_template.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        pd.DataFrame(
            self.oversize_rate_table.get_rows(),
            columns=self.oversize_rate_table.columns
        ).to_excel(path, index=False)

        messagebox.showinfo(
            "Done",
            f"Oversize template exported:\n{path}"
        )

    def import_oversize_rate_template(self):
        def validator(df: pd.DataFrame):
            rows = []
            seen = set()
            for i, row in df.iterrows():
                excel_row = i + 2
                service = str(row["Service"]).strip()

                if service not in SERVICES:
                    self._raise_template_error("Oversize", f"Row {excel_row}: Service must be one of {SERVICES}")

                try:
                    zone = int(float(row["Zone"]))
                except Exception:
                    self._raise_template_error("Oversize", f"Row {excel_row}: Zone must be numeric")
                if zone not in ZONES:
                    self._raise_template_error("Oversize", f"Row {excel_row}: Zone must be one of {ZONES}")

                try:
                    fee = self._normalize_numeric_cell(row["Fee"], 0.0)
                except Exception:
                    self._raise_template_error("Oversize", f"Row {excel_row}: Fee must be numeric")

                key = (service, zone)
                if key in seen:
                    self._raise_template_error("Oversize", f"Row {excel_row}: Duplicate Service + Zone")
                seen.add(key)

                rows.append([service, str(zone), str(fee)])

            if not rows:
                self._raise_template_error("Oversize", "Template is empty.")
            return rows

        self._import_table_with_validation(
            self.oversize_rate_table,
            "Oversize",
            ["Service", "Zone", "Fee"],
            validator,
        )

    def export_residential_rate_template(self):

        path = filedialog.asksaveasfilename(
            initialfile="Residential_template.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        pd.DataFrame(
            self.residential_rate_table.get_rows(),
            columns=self.residential_rate_table.columns
        ).to_excel(path, index=False)

        messagebox.showinfo(
            "Done",
            f"Residential template exported:\n{path}"
        )

    def import_residential_rate_template(self):
        def validator(df: pd.DataFrame):
            rows = []
            seen = set()
            for i, row in df.iterrows():
                excel_row = i + 2
                service = str(row["Service"]).strip()

                if service not in SERVICES:
                    self._raise_template_error("Residential", f"Row {excel_row}: Service must be one of {SERVICES}")

                try:
                    fee = self._normalize_numeric_cell(row["Fee"], 0.0)
                except Exception:
                    self._raise_template_error("Residential", f"Row {excel_row}: Fee must be numeric")

                if service in seen:
                    self._raise_template_error("Residential", f"Row {excel_row}: Duplicate Service {service}")
                seen.add(service)

                rows.append([service, str(fee)])

            if not rows:
                self._raise_template_error("Residential", "Template is empty.")
            return rows

        self._import_table_with_validation(
            self.residential_rate_table,
            "Residential",
            ["Service", "Fee"],
            validator,
        )

    def export_signature_rate_template(self):

        path = filedialog.asksaveasfilename(
            initialfile="Signature_template.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        pd.DataFrame(
            self.signature_rate_table.get_rows(),
            columns=self.signature_rate_table.columns
        ).to_excel(path, index=False)

        messagebox.showinfo(
            "Done",
            f"Signature template exported:\n{path}"
        )

    def import_signature_rate_template(self):
        def validator(df: pd.DataFrame):
            rows = []
            seen = set()
            for i, row in df.iterrows():
                excel_row = i + 2
                sig_type = str(row["Signature Type"]).strip()

                if sig_type not in SIGNATURE_TYPES:
                    self._raise_template_error("Signature", f"Row {excel_row}: Signature Type must be one of {SIGNATURE_TYPES}")

                try:
                    fee = self._normalize_numeric_cell(row["Fee"], 0.0)
                except Exception:
                    self._raise_template_error("Signature", f"Row {excel_row}: Fee must be numeric")

                if sig_type in seen:
                    self._raise_template_error("Signature", f"Row {excel_row}: Duplicate Signature Type {sig_type}")
                seen.add(sig_type)

                rows.append([sig_type, str(fee)])

            if not rows:
                self._raise_template_error("Signature", "Template is empty.")
            return rows

        self._import_table_with_validation(
            self.signature_rate_table,
            "Signature",
            ["Signature Type", "Fee"],
            validator,
        )

    def export_general_demand_template(self):
        self._export_table(self.general_demand_table, "General Demand")

    def import_general_demand_template(self):
        self._import_table_with_validation(
            self.general_demand_table,
            "General Demand",
            ["Service", "Start Date", "End Date", "Fee"],
            lambda df: self._validate_demand_template("General Demand", df),
        )

    def export_ahs_demand_template(self):
        self._export_table(self.ahs_demand_table, "AHS Demand")

    def import_ahs_demand_template(self):
        self._import_table_with_validation(
            self.ahs_demand_table,
            "AHS Demand",
            ["Service", "Start Date", "End Date", "Fee"],
            lambda df: self._validate_demand_template("AHS Demand", df),
        )

    def export_oversize_demand_template(self):
        self._export_table(self.oversize_demand_table, "Oversize Demand")

    def import_oversize_demand_template(self):
        self._import_table_with_validation(
            self.oversize_demand_table,
            "Oversize Demand",
            ["Service", "Start Date", "End Date", "Fee"],
            lambda df: self._validate_demand_template("Oversize Demand", df),
        )

    def export_billing_sample(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        df = pd.DataFrame([
            {
                "Tracking Number": "TEST001",
                "Service": "Ground",
                "Zone": 2,
                "Length": 20,
                "Width": 14,
                "Height": 10,
                "Actual Weight": 12,
                "Residential": "Y",
                "DAS Type": "DAS Comm",
                "Adult Signature": 10.00,
                "Address Correction": 6.38,
                "Shipment Date": "2026-03-26",
            },
            {
                "Tracking Number": "TEST002",
                "Service": "Home Delivery",
                "Zone": 5,
                "Length": 32,
                "Width": 20,
                "Height": 18,
                "Actual Weight": 22,
                "Residential": "Y",
                "DAS Type": "DAS Extended Resi",
                "Direct Signature": 5.70,
                "Shipment Date": "2026-03-27",
            },
        ])
        df.to_excel(path, index=False)
        messagebox.showinfo("Done", f"Billing sample exported:\n{path}")

    def save_config(self):
        cfg = {
            "global_rules": {
                "dim_factor": self.dim_factor.get(),
                "fuel_percent": self.fuel_percent.get(),
                "use_oversize_longest": self.use_oversize_longest.get(),
                "use_oversize_weight": self.use_oversize_weight.get(),
                "use_oversize_cubic": self.use_oversize_cubic.get(),
                "use_oversize_lg": self.use_oversize_lg.get(),
                "oversize_longest_side": self.oversize_longest_side.get(),
                "oversize_actual_weight": self.oversize_actual_weight.get(),
                "oversize_cubic_inches": self.oversize_cubic_inches.get(),
                "length_girth_limit": self.length_girth_limit.get(),
                "oversize_min_billable_weight": self.oversize_min_billable_weight.get(),
                "ahs_weight_threshold": self.ahs_weight_threshold.get(),
                "ahs_dimension_lg_limit": self.ahs_dimension_lg_limit.get(),
                "ahs_dimension_longest_side": self.ahs_dimension_longest_side.get(),
                "ahs_dimension_second_side": self.ahs_dimension_second_side.get(),
                "ahs_dimension_cubic_inches": self.ahs_dimension_cubic_inches.get(),
                "ahs_min_billable_weight": self.ahs_min_billable_weight.get(),
            },
            "base_rate_rows": self.base_rate_table.get_rows(),
            "ahs_rate_rows": self.ahs_rate_table.get_rows(),
            "das_rate_rows": self.das_rate_table.get_rows(),
            "oversize_rate_rows": self.oversize_rate_table.get_rows(),
            "residential_rate_rows": self.residential_rate_table.get_rows(),
            "signature_rate_rows": self.signature_rate_table.get_rows(),
            "general_demand_rows": self.general_demand_table.get_rows(),
            "ahs_demand_rows": self.ahs_demand_table.get_rows(),
            "oversize_demand_rows": self.oversize_demand_table.get_rows(),
        }

        path = self.config_path.get().strip() or DEFAULT_CONFIG_PATH
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)

        messagebox.showinfo("Done", f"Config saved:\n{path}")

    def load_config(self):
        path = self.config_path.get().strip()
        if not path or not os.path.exists(path):
            raise UserInputError("Config file not found.")

        with open(path, "r", encoding="utf-8") as f:
            cfg = json.load(f)

        rules = cfg.get("global_rules", {})
        mappings = [
            ("dim_factor", self.dim_factor),
            ("fuel_percent", self.fuel_percent),
            ("use_oversize_longest", self.use_oversize_longest),
            ("use_oversize_weight", self.use_oversize_weight),
            ("use_oversize_cubic", self.use_oversize_cubic),
            ("use_oversize_lg", self.use_oversize_lg),
            ("oversize_longest_side", self.oversize_longest_side),
            ("oversize_actual_weight", self.oversize_actual_weight),
            ("oversize_cubic_inches", self.oversize_cubic_inches),
            ("length_girth_limit", self.length_girth_limit),
            ("oversize_min_billable_weight", self.oversize_min_billable_weight),
            ("ahs_weight_threshold", self.ahs_weight_threshold),
            ("ahs_dimension_lg_limit", self.ahs_dimension_lg_limit),
            ("ahs_dimension_longest_side", self.ahs_dimension_longest_side),
            ("ahs_dimension_second_side", self.ahs_dimension_second_side),
            ("ahs_dimension_cubic_inches", self.ahs_dimension_cubic_inches),
            ("ahs_min_billable_weight", self.ahs_min_billable_weight),
        ]
        for key, var in mappings:
            if key in rules:
                var.set(rules[key])

        # Backward compatibility for old config files created before AHS Dimension was corrected.
        # Old keys named these as Package thresholds, but they were actually Dimension triggers.
        if "ahs_dimension_longest_side" not in rules and "ahs_package_longest" in rules:
            self.ahs_dimension_longest_side.set(rules["ahs_package_longest"])
        if "ahs_dimension_second_side" not in rules and "ahs_package_second" in rules:
            self.ahs_dimension_second_side.set(rules["ahs_package_second"])

        self.base_rate_table.load_rows(cfg.get("base_rate_rows", []))
        self.ahs_rate_table.load_rows(cfg.get("ahs_rate_rows", []))
        self.das_rate_table.load_rows(cfg.get("das_rate_rows", []))
        self.oversize_rate_table.load_rows(cfg.get("oversize_rate_rows", []))
        self.residential_rate_table.load_rows(cfg.get("residential_rate_rows", []))
        self.signature_rate_table.load_rows(cfg.get("signature_rate_rows", []))
        self.general_demand_table.load_rows(cfg.get("general_demand_rows", []))
        self.ahs_demand_table.load_rows(cfg.get("ahs_demand_rows", []))
        self.oversize_demand_table.load_rows(cfg.get("oversize_demand_rows", []))

        messagebox.showinfo("Done", f"Config loaded:\n{path}")

    def _read_billing(self) -> pd.DataFrame:
        path = self.billing_path.get().strip()
        if not path:
            raise UserInputError("Please choose a billing file first.")

        if not os.path.exists(path):
            raise UserInputError(f"Billing file not found:\n{path}")

        try:
            if path.lower().endswith(".csv"):
                return pd.read_csv(path)
            return pd.read_excel(path)
        except Exception as e:
            raise UserInputError(f"Failed to read billing file.\n\n{e}")

    def _coalesce(self, row: pd.Series, candidates: List[str], default=""):
        for col in candidates:
            if col in row.index and pd.notna(row[col]) and str(row[col]).strip() != "":
                return row[col]
        return default

    def _has_nonzero(self, row: pd.Series, col: str) -> bool:
        if col not in row.index:
            return False
        return abs(self._to_float(row[col], 0.0)) > 0

    def _parse_date(self, value) -> Optional[datetime]:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return None
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, datetime):
            return value

        s = str(value).strip()
        if not s:
            return None

        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]

        for fmt in [
            "%Y%m%d",
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%m-%d-%Y",
        ]:
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass

        try:
            return pd.to_datetime(s, errors="raise").to_pydatetime()
        except Exception:
            return None

    def _get_effective_date(self, row: pd.Series) -> Optional[datetime]:
        for col in ["Shipment Date", "Tendered Date", "Invoice Date"]:
            if col in row.index:
                dt = self._parse_date(row[col])
                if dt:
                    return dt
        return None

    def _match_demand_fee(self, table: RateTableFrame, service: str, effective_date: Optional[datetime]) -> float:
        if not effective_date:
            return 0.0

        matched_fees = []
        for row in table.get_rows():
            if len(row) < 4:
                continue

            row_service = str(row[0]).strip()
            start_date = self._parse_date(row[1])
            end_date = self._parse_date(row[2])
            fee = self._to_float(row[3], 0.0)

            if row_service != service:
                continue
            if not start_date or not end_date:
                continue

            if start_date <= effective_date <= end_date:
                matched_fees.append(fee)

        return max(matched_fees) if matched_fees else 0.0

    def _iter_charge_description_amount_pairs(self, row: pd.Series) -> List[Tuple[str, float]]:
        pairs = []
        desc_prefixes = [
            "Tracking ID Charge Description",
            "Charge Description",
        ]

        for col in row.index:
            col_name = str(col).strip()
            matched_prefix = None
            for prefix in desc_prefixes:
                if col_name == prefix or col_name.startswith(prefix + "."):
                    matched_prefix = prefix
                    break

            if not matched_prefix:
                continue

            suffix = col_name[len(matched_prefix):]
            amount_col = matched_prefix.replace("Description", "Amount") + suffix

            desc = "" if pd.isna(row[col]) else str(row[col]).strip()
            if not desc:
                continue

            amount = 0.0
            if amount_col in row.index:
                amount = self._to_float(row[amount_col], 0.0)

            pairs.append((desc, amount))

        return pairs

    @staticmethod
    def _norm_charge_desc(desc: str) -> str:
        return " ".join(str(desc).lower().replace("’", "'").replace("–", "-").split())

    def _has_charge_desc(self, row: pd.Series, keywords: List[str], require_nonzero: bool = True) -> bool:
        normalized_keywords = [self._norm_charge_desc(k) for k in keywords]
        for desc, amount in self._iter_charge_description_amount_pairs(row):
            d = self._norm_charge_desc(desc)
            if any(k in d for k in normalized_keywords):
                if not require_nonzero or abs(amount) > 0:
                    return True
        return False

    def _charge_desc_amount(self, row: pd.Series, keywords: List[str]) -> float:
        normalized_keywords = [self._norm_charge_desc(k) for k in keywords]
        total = 0.0
        for desc, amount in self._iter_charge_description_amount_pairs(row):
            d = self._norm_charge_desc(desc)
            if any(k in d for k in normalized_keywords):
                total += amount
        return total

    def _detect_das_type_from_raw_invoice(self, row: pd.Series) -> str:
        paired_mapping = [
            (["DAS Remote Residential"], "DAS Remote Residential"),
            (["DAS Extended Residential"], "DAS Extended Resi"),
            (["DAS Residential"], "DAS Resi"),
            (["DAS Remote Commercial"], "DAS Remote Commercial"),
            (["DAS Extended Commercial"], "DAS Extended Comm"),
            (["DAS Commercial"], "DAS Comm"),
            (["Hawaii - Residential", "Hawaii Residential"], "Hawaii - Residential"),
            (["Hawaii - Commercial", "Hawaii Commercial"], "Hawaii - Commercial"),
            (["Alaska - Residential", "Alaska Residential"], "Alaska - Residential"),
            (["Alaska - Commercial", "Alaska Commercial"], "Alaska - Commercial"),
        ]
        for keywords, das_type in paired_mapping:
            if self._has_charge_desc(row, keywords):
                return das_type

        standalone_mapping = [
            ("DAS Remote Residential", "DAS Remote Residential"),
            ("DAS Extended Residential", "DAS Extended Resi"),
            ("DAS Residential", "DAS Resi"),
            ("DAS Remote Commercial", "DAS Remote Commercial"),
            ("DAS Extended Commercial", "DAS Extended Comm"),
            ("DAS Commercial", "DAS Comm"),
            ("Hawaii - Commercial", "Hawaii - Commercial"),
            ("Hawaii - Residential", "Hawaii - Residential"),
            ("Alaska - Commercial", "Alaska - Commercial"),
            ("Alaska - Residential", "Alaska - Residential"),
        ]
        for raw_col, das_type in standalone_mapping:
            if self._has_nonzero(row, raw_col):
                return das_type

        raw_das = str(self._coalesce(row, ["DAS Type"], "")).strip()
        if raw_das in DAS_TYPES:
            return raw_das
        return ""

    def _detect_residential_flag_from_raw_invoice(self, row: pd.Series) -> bool:
        if self._flag(self._coalesce(row, ["Residential"], "")):
            return True

        if self._has_charge_desc(
            row,
            [
                "Residential Delivery",
                "Residential Surcharge",
                "Residential",
                "DAS Residential",
                "DAS Extended Residential",
                "DAS Remote Residential",
                "Hawaii - Residential",
                "Alaska - Residential",
            ],
        ):
            return True

        residential_charge_cols = [
            "Residential Delivery",
            "Residential Surcharge",
            "DAS Residential",
            "DAS Extended Residential",
            "DAS Remote Residential",
            "Hawaii - Residential",
            "Alaska - Residential",
        ]
        for col in residential_charge_cols:
            if self._has_nonzero(row, col):
                return True
        return False

    def _detect_signature_type_from_row(self, row: pd.Series) -> str:
        if self._has_charge_desc(row, ["Adult Signature", "Adult Signature Req"]):
            return "Adult Signature"
        if self._has_charge_desc(row, ["Direct Signature", "Direct Signature Req"]):
            return "Direct Signature"
        if self._has_charge_desc(row, ["Indirect Signature", "Indirect Signature Req"]):
            return "Indirect Signature"

        if self._has_nonzero(row, "Adult Signature") or self._has_nonzero(row, "Adult Signature Req."):
            return "Adult Signature"
        if self._has_nonzero(row, "Direct Signature") or self._has_nonzero(row, "Direct Signature Req."):
            return "Direct Signature"
        if self._has_nonzero(row, "Indirect Signature") or self._has_nonzero(row, "Indirect Signature Req."):
            return "Indirect Signature"

        raw_sig = str(self._coalesce(row, ["Signature Type"], "")).strip()
        if raw_sig in SIGNATURE_TYPES:
            return raw_sig
        return ""

    def _detect_ahs_result_from_row(
        self,
        row: pd.Series,
        l: float,
        w: float,
        h: float,
        actual_weight: float,
        oversize_hit: bool,
        service: str,
        zone: int,
        ahs_lookup: Dict[Tuple[str, str, int], float],
    ) -> Dict[str, object]:
        """Return final AHS result using highest-only logic.

        Important business rules:
        - Packaging is detected only from invoice charge descriptions / raw charge columns.
        - Dimension is detected from thresholds or explicit invoice dimension charge.
        - If multiple AHS types hit, only the highest fee is charged.
        - Oversize shipment suppresses AHS.
        """
        if oversize_hit:
            return {
                "ahs_type": "",
                "ahs_fee": 0.0,
                "dimension_triggered": False,
            }

        dims = sorted([l, w, h], reverse=True)
        longest = dims[0]
        second_longest = dims[1]
        shortest = dims[2]
        cubic = l * w * h
        length_plus_girth = longest + 2 * (second_longest + shortest)

        ahs_candidates = []

        packaging_detected = (
            self._has_charge_desc(
                row,
                [
                    "Add'l Handling-Packaging",
                    "Additional Handling-Packaging",
                    "AHS - Packaging",
                ],
            )
            or any(
                self._has_nonzero(row, col)
                for col in [
                    "Add'l Handling-Packaging",
                    "Add'l Handling - Packaging",
                    "Additional Handling-Packaging",
                    "Additional Handling - Packaging",
                    "AHS - Packaging",
                    "AHS Packaging",
                ]
            )
        )

        if packaging_detected:
            ahs_candidates.append({
                "type": "Package",
                "fee": ahs_lookup.get((service, "Package", zone), 0.0),
            })

        if actual_weight > self.ahs_weight_threshold.get():
            ahs_candidates.append({
                "type": "Weight",
                "fee": ahs_lookup.get((service, "Weight", zone), 0.0),
            })

        dimension_triggered = (
            length_plus_girth > self.ahs_dimension_lg_limit.get()
            or cubic > self.ahs_dimension_cubic_inches.get()
            or longest > self.ahs_dimension_longest_side.get()
            or second_longest > self.ahs_dimension_second_side.get()
            or self._has_charge_desc(
                row,
                [
                    "Add'l Handling-Dimension",
                    "Additional Handling-Dimension",
                    "AHS - Dimension",
                    "AHS - Dimensions",
                ],
            )
            or any(
                self._has_nonzero(row, col)
                for col in [
                    "Add'l Handling-Dimension",
                    "Additional Handling-Dimension",
                    "AHS - Dimension",
                    "AHS - Dimensions",
                ]
            )
        )

        if dimension_triggered:
            ahs_candidates.append({
                "type": "Dimension",
                "fee": ahs_lookup.get((service, "Dimension", zone), 0.0),
            })

        if not ahs_candidates:
            return {
                "ahs_type": "",
                "ahs_fee": 0.0,
                "dimension_triggered": False,
            }

        highest = max(ahs_candidates, key=lambda x: float(x["fee"]))
        return {
            "ahs_type": highest["type"],
            "ahs_fee": float(highest["fee"]),
            "dimension_triggered": bool(dimension_triggered),
        }

    def _normalize_billing_input(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]

        simplified_required = ["Service", "Zone", "Length", "Width", "Height", "Actual Weight"]
        if all(col in df.columns for col in simplified_required):
            out = df.copy()

            if "Residential" not in out.columns:
                out["Residential"] = ""
            if "DAS Type" not in out.columns:
                out["DAS Type"] = ""
            if "Tracking Number" not in out.columns:
                out["Tracking Number"] = ""

            out["_Tracking Number"] = out["Tracking Number"].apply(lambda x: "" if pd.isna(x) else str(x))
            out["_Normalized Service"] = out["Service"].apply(lambda x: self._normalize_service(x))
            out["_Normalized Zone"] = out["Zone"].apply(lambda x: int(self._to_float(x, 0)))
            out["_Normalized Length"] = out["Length"].apply(lambda x: self._to_float(x, 0.0))
            out["_Normalized Width"] = out["Width"].apply(lambda x: self._to_float(x, 0.0))
            out["_Normalized Height"] = out["Height"].apply(lambda x: self._to_float(x, 0.0))
            out["_Normalized Actual Weight"] = out["Actual Weight"].apply(lambda x: self._to_float(x, 0.0))
            out["_Normalized Residential"] = out["Residential"].apply(lambda x: "Y" if self._flag(x) else "N")
            out["_Normalized DAS Type"] = out["DAS Type"].fillna("").astype(str)
            out["_Normalized Signature Type"] = ""
            return out

        raw_required = ["Zone Code", "Dim Length", "Dim Width", "Dim Height", "Actual Weight Amount"]
        if not all(col in df.columns for col in raw_required):
            raise UserInputError(
                "Unsupported billing format.\n\n"
                "Need either simplified columns:\n"
                "Service, Zone, Length, Width, Height, Actual Weight\n\n"
                "or raw FedEx invoice columns like:\n"
                "Zone Code, Dim Length, Dim Width, Dim Height, Actual Weight Amount"
            )

        out = df.copy()

        out["_Tracking Number"] = pd.Series([""] * len(out), dtype="object")
        out["_Normalized Service"] = pd.Series([""] * len(out), dtype="object")
        out["_Normalized Zone"] = pd.Series([0] * len(out), dtype="object")
        out["_Normalized Length"] = pd.Series([0.0] * len(out), dtype="object")
        out["_Normalized Width"] = pd.Series([0.0] * len(out), dtype="object")
        out["_Normalized Height"] = pd.Series([0.0] * len(out), dtype="object")
        out["_Normalized Actual Weight"] = pd.Series([0.0] * len(out), dtype="object")
        out["_Normalized Residential"] = pd.Series([""] * len(out), dtype="object")
        out["_Normalized DAS Type"] = pd.Series([""] * len(out), dtype="object")
        out["_Normalized Signature Type"] = pd.Series([""] * len(out), dtype="object")

        for idx, row in out.iterrows():
            try:
                service_value = self._coalesce(row, ["Service Type", "Ground Service", "Service"], "")
                service = self._normalize_service(service_value)

                zone = int(self._to_float(self._coalesce(row, ["Zone Code", "Zone"], 0), 0))
                length = self._to_float(self._coalesce(row, ["Dim Length", "Length"], 0), 0)
                width = self._to_float(self._coalesce(row, ["Dim Width", "Width"], 0), 0)
                height = self._to_float(self._coalesce(row, ["Dim Height", "Height"], 0), 0)
                actual_weight = self._to_float(self._coalesce(row, ["Actual Weight Amount", "Rated Weight Amount", "Actual Weight"], 0), 0)

                tracking = self._coalesce(
                    row,
                    ["Express or Ground Tracking ID", "CrossRefTrackingID", "Tracking Number"],
                    ""
                )

                out.at[idx, "_Tracking Number"] = "" if pd.isna(tracking) else str(tracking)
                out.at[idx, "_Normalized Service"] = str(service)
                out.at[idx, "_Normalized Zone"] = zone
                out.at[idx, "_Normalized Length"] = length
                out.at[idx, "_Normalized Width"] = width
                out.at[idx, "_Normalized Height"] = height
                out.at[idx, "_Normalized Actual Weight"] = actual_weight
                out.at[idx, "_Normalized Residential"] = "Y" if self._detect_residential_flag_from_raw_invoice(row) else "N"
                out.at[idx, "_Normalized DAS Type"] = str(self._detect_das_type_from_raw_invoice(row))
                out.at[idx, "_Normalized Signature Type"] = str(self._detect_signature_type_from_row(row))

            except Exception as e:
                raise UserInputError(f"Failed to normalize billing row {idx + 2}.\n\n{e}")

        return out

    def _base_lookup(self) -> Dict[Tuple[str, int, int], float]:
        lookup = {}
        for row in self.base_rate_table.get_rows():
            service = str(row[0]).strip()
            weight = int(self._to_float(row[1], 0))
            if not service or weight <= 0:
                continue
            for idx, zone in enumerate(ZONES, start=2):
                lookup[(service, weight, zone)] = self._to_float(row[idx], 0)
        return lookup

    def _ahs_lookup(self):

        lookup = {}

        for row in self.ahs_rate_table.get_rows():

            service = str(row[0]).strip()
            ahs_type = str(row[1]).strip()
            zone = int(self._to_float(row[2], 0))

            fee = self._to_float(row[3], 0)

            lookup[(service, ahs_type, zone)] = fee

        return lookup

    def _das_lookup(self) -> Dict[Tuple[str, str, int], float]:
        lookup = {}
        for row in self.das_rate_table.get_rows():
            service = str(row[0]).strip()
            das_type = str(row[1]).strip()
            zone = int(self._to_float(row[2], 0))
            fee = self._to_float(row[3], 0)
            lookup[(service, das_type, zone)] = fee
        return lookup

    def _oversize_lookup(self) -> Dict[Tuple[str, int], float]:
        lookup = {}
        for row in self.oversize_rate_table.get_rows():
            service = str(row[0]).strip()
            zone = int(self._to_float(row[1], 0))
            fee = self._to_float(row[2], 0)
            lookup[(service, zone)] = fee
        return lookup

    def _residential_lookup(self) -> Dict[str, float]:
        lookup = {}
        for row in self.residential_rate_table.get_rows():
            service = str(row[0]).strip()
            fee = self._to_float(row[1], 0)
            if service:
                lookup[service] = fee
        return lookup

    def _signature_lookup(self) -> Dict[str, float]:
        lookup = {}
        for row in self.signature_rate_table.get_rows():
            sig_type = str(row[0]).strip()
            fee = self._to_float(row[1], 0)
            if sig_type:
                lookup[sig_type] = fee
        return lookup

    def _calc_dim_weight(self, l: float, w: float, h: float) -> int:
        return math.ceil((l * w * h) / max(self.dim_factor.get(), 1.0))

    def _oversize_hit(self, l: float, w: float, h: float, actual_weight: float) -> bool:
        longest = max(l, w, h)
        cubic = l * w * h
        girth = 2 * (w + h)
        length_plus_girth = longest + girth

        conditions = []

        if self.use_oversize_longest.get():
            conditions.append(longest > self.oversize_longest_side.get())
        if self.use_oversize_weight.get():
            conditions.append(actual_weight > self.oversize_actual_weight.get())
        if self.use_oversize_cubic.get():
            conditions.append(cubic > self.oversize_cubic_inches.get())
        if self.use_oversize_lg.get():
            conditions.append(length_plus_girth > self.length_girth_limit.get())

        return any(conditions)

    def _get_base_rate_or_raise(self, base_lookup: Dict[Tuple[str, int, int], float], service: str, weight: int, zone: int) -> float:
        if weight > BASE_RATE_MAX_WEIGHT:
            raise UserInputError(
                f"Billable weight {weight} exceeds base rate limit {BASE_RATE_MAX_WEIGHT}.\n"
                f"Service={service}\nZone={zone}"
            )
        key = (service, weight, zone)
        if key not in base_lookup:
            raise UserInputError(
                f"Missing base rate.\n\nService={service}\nWeight={weight}\nZone={zone}\n\n"
                f"Please check Base Rate template."
            )
        return base_lookup[key]

    def _get_oversize_fee_or_raise(self, oversize_lookup: Dict[Tuple[str, int], float], service: str, zone: int) -> float:
        key = (service, zone)
        if key not in oversize_lookup:
            raise UserInputError(
                f"Missing oversize fee.\n\nService={service}\nZone={zone}\n\n"
                f"Please check Oversize template."
            )
        return oversize_lookup[key]

    def _collect_other_pass_through_fees(self, row: pd.Series) -> Dict[str, float]:
        def is_excluded_charge_name(name: str) -> bool:
            d = self._norm_charge_desc(name)
            excluded_keywords = [
                "fuel surcharge",
                "fuel",
                "variable handling charge",
                "delivery area surcharge",
                "das residential",
                "das extended residential",
                "das remote residential",
                "das commercial",
                "das extended commercial",
                "das remote commercial",
                "hawaii - commercial",
                "hawaii - residential",
                "alaska - commercial",
                "alaska - residential",
                "residential delivery",
                "residential surcharge",
                "add'l handling",
                "additional handling",
                "ahs",
                "oversize charge",
                "signature",
                "transportation charge",
                "net charge",
                "discount",
                "earned discount",
                "performance pricing",
            ]
            return any(k in d for k in excluded_keywords)

        pass_through = {}

        for desc, amount in self._iter_charge_description_amount_pairs(row):
            if abs(amount) <= 0:
                continue
            if is_excluded_charge_name(desc):
                continue
            pass_through[str(desc)] = pass_through.get(str(desc), 0.0) + amount

        excluded_cols = {
            "Zone Code", "Dim Length", "Dim Width", "Dim Height", "Actual Weight Amount",
            "Length", "Width", "Height", "Actual Weight", "Zone",
            "Fuel Surcharge", "Fuel", "Net Fuel Surcharge", "Variable Handling Charge",
            "DAS Residential", "DAS Extended Residential", "DAS Remote Residential",
            "DAS Commercial", "DAS Extended Commercial", "DAS Remote Commercial",
            "Hawaii - Commercial", "Hawaii - Residential",
            "Alaska - Commercial", "Alaska - Residential",
            "Residential Delivery", "Residential Surcharge", "Residential",
            "Add'l Handling-Weight", "Add'l Handling-Dimension", "Add'l Handling-Packaging", "AHS - Dimensions",
            "Oversize Charge",
            "Adult Signature", "Direct Signature", "Indirect Signature",
            "Adult Signature Req.", "Direct Signature Req.", "Indirect Signature Req.",
            "Transportation Charge Amount", "Net Charge Amount",
            "Discount", "Earned Discount", "Performance Pricing",
        }

        for col in row.index:
            col_str = str(col).strip()
            col_lower = col_str.lower()
            if col_str.startswith("_") or col_str.startswith("CALC_"):
                continue
            if col_str in excluded_cols:
                continue
            if col_lower == "tracking id charge description" or col_lower.startswith("tracking id charge description."):
                continue
            if col_lower == "tracking id charge amount" or col_lower.startswith("tracking id charge amount."):
                continue
            if col_lower == "charge description" or col_lower.startswith("charge description."):
                continue
            if col_lower == "charge amount" or col_lower.startswith("charge amount."):
                continue
            if is_excluded_charge_name(col_str):
                continue

            likely_fee = any(
                kw in col_lower
                for kw in ["charge", "surcharge", "billing", "correction", "unauthorized", "third party", "handling"]
            )
            if not likely_fee:
                continue

            fee = self._to_float(row[col], None)
            if fee is None:
                continue

            if abs(fee) > 0:
                pass_through[col_str] = pass_through.get(col_str, 0.0) + fee

        return pass_through

    def _apply_excel_formatting(self, output_path: str):
        wb = load_workbook(output_path)
        ws = wb["Repriced"]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        currency_cols = {
            "Base Rate", "Additional Handling Surcharge", "DAS Fee", "Oversize Fee", "Residential Fee",
            "Signature Fee", "General Demand Fee", "AHS Demand Fee", "Oversize Demand Fee",
            "Fuel Fee", "Repriced Total"
        }
        date_cols = {"Transaction Date", "Invoice Date"}
        numeric_cols = {
            "Length", "Width", "Height", "Actual Weight", "DIM Weight", "Billable Weight"
        }
        percent_cols = set()

        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            if header in currency_cols:
                fmt = '$#,##0.00'
            elif header in date_cols:
                fmt = 'mm/dd/yyyy'
            elif header in numeric_cols:
                fmt = '0.00'
            elif header in percent_cols:
                fmt = '0.00%'
            else:
                fmt = None

            if fmt:
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = fmt

        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 35)

        wb.save(output_path)

    def run_repricing(self):
        try:
            raw_billing = self._read_billing()
            billing = self._normalize_billing_input(raw_billing)

            base_lookup = self._base_lookup()
            ahs_lookup = self._ahs_lookup()
            das_lookup = self._das_lookup()
            oversize_lookup = self._oversize_lookup()
            residential_lookup = self._residential_lookup()
            signature_lookup = self._signature_lookup()

            for item in self.preview.get_children():
                self.preview.delete(item)

            output = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Save Repriced Result As",
                initialfile="repriced_result.xlsx"
            )

            if not output:
                return

            result_rows = []

            for index, row in billing.iterrows():
                excel_row = index + 2

                service = self._normalize_service(row["_Normalized Service"])
                zone = int(self._to_float(row["_Normalized Zone"], 0))
                l = self._to_float(row["_Normalized Length"], 0)
                w = self._to_float(row["_Normalized Width"], 0)
                h = self._to_float(row["_Normalized Height"], 0)
                actual_weight = math.ceil(self._to_float(row["_Normalized Actual Weight"], 0))

                if zone not in ZONES:
                    print(f"WARNING: Billing row {excel_row}: Zone {zone} not configured")
                    continue

                effective_date = self._get_effective_date(row)

                dim_weight = self._calc_dim_weight(l, w, h)

                rated_weight = self._to_float(
                    self._coalesce(row, ["Rated Weight Amount", "Billed Weight", "Billable Weight"], 0),
                    0
                )

                if rated_weight > 0:
                    billable_weight = math.ceil(rated_weight)
                else:
                    billable_weight = max(actual_weight, dim_weight)

                oversize_hit = self._oversize_hit(l, w, h, actual_weight)
                if "Oversize Charge" in billing.columns and self._to_float(row.get("Oversize Charge", 0.0), 0.0) > 0:
                    oversize_hit = True

                ahs_result = self._detect_ahs_result_from_row(
                    row,
                    l,
                    w,
                    h,
                    actual_weight,
                    oversize_hit,
                    service,
                    zone,
                    ahs_lookup,
                )
                ahs_type = str(ahs_result["ahs_type"])
                ahs_fee = float(ahs_result["ahs_fee"])
                dimension_triggered = bool(ahs_result["dimension_triggered"])

                if oversize_hit:
                    billable_weight = max(billable_weight, math.ceil(self.oversize_min_billable_weight.get()))
                elif dimension_triggered:
                    # FedEx AHS minimum billable weight applies to dimensional AHS triggers
                    # such as Length + Girth, longest side, second-longest side, or cubic size.
                    # AHS Weight and AHS Packaging do not use this AHS minimum billable weight.
                    billable_weight = max(billable_weight, math.ceil(self.ahs_min_billable_weight.get()))

                base_rate = self._get_base_rate_or_raise(base_lookup, service, billable_weight, zone)


                das_type = str(row.get("_Normalized DAS Type", "")).strip()
                if das_type and das_type not in DAS_TYPES:
                    raise UserInputError(f"Billing row {excel_row}: DAS Type '{das_type}' is invalid.")
                das_fee = 0.0 if not das_type else das_lookup.get((service, das_type, zone), 0.0)

                oversize_fee = self._get_oversize_fee_or_raise(oversize_lookup, service, zone) if oversize_hit else 0.0

                residential_fee = 0.0
                if self._flag(row.get("_Normalized Residential", "")):
                    residential_fee = residential_lookup.get(service, 0.0)

                signature_type = str(row.get("_Normalized Signature Type", "")).strip()
                if signature_type and signature_type not in SIGNATURE_TYPES:
                    raise UserInputError(f"Billing row {excel_row}: Signature Type '{signature_type}' is invalid.")
                signature_fee = 0.0 if not signature_type else signature_lookup.get(signature_type, 0.0)

                general_demand_fee = self._match_demand_fee(self.general_demand_table, service, effective_date)
                ahs_demand_fee = self._match_demand_fee(self.ahs_demand_table, service, effective_date) if ahs_fee > 0 else 0.0
                oversize_demand_fee = self._match_demand_fee(self.oversize_demand_table, service, effective_date) if oversize_fee > 0 else 0.0

                other_pass = self._collect_other_pass_through_fees(row)
                other_pass_total = sum(other_pass.values())

                total_demand_fee = general_demand_fee + ahs_demand_fee + oversize_demand_fee

                fuel_base = (
                    base_rate
                    + ahs_fee
                    + das_fee
                    + oversize_fee
                    + residential_fee
                    + signature_fee
                    + total_demand_fee
                    + other_pass_total
                )
                fuel_fee = fuel_base * self.fuel_percent.get() / 100.0
                total = fuel_base + fuel_fee

                transaction_date = self._coalesce(row, ["Shipment Date", "Transaction Date", "Tendered Date"], "")
                invoice_date = self._coalesce(row, ["Invoice Date"], "")
                invoice_number = self._coalesce(row, ["Invoice Number", "Invoice No", "Invoice #"], "")

                out = row.to_dict()
                out.update({
                    "Transaction Date": transaction_date,
                    "Invoice Date": invoice_date,
                    "Invoice Number": invoice_number,
                    "Tracking Number": row.get("_Tracking Number", ""),
                    "Service": service,
                    "Zone": zone,
                    "Length": l,
                    "Width": w,
                    "Height": h,
                    "Actual Weight": actual_weight,
                    "DIM Weight": dim_weight,
                    "Billable Weight": billable_weight,

                    "Base Rate": round(base_rate, 2),
                    "Additional Handling Surcharge": round(ahs_fee, 2),
                    "DAS Fee": round(das_fee, 2),
                    "Oversize Fee": round(oversize_fee, 2),
                    "Residential Fee": round(residential_fee, 2),
                    "Signature Fee": round(signature_fee, 2),
                    "General Demand Fee": round(general_demand_fee, 2),
                    "AHS Demand Fee": round(ahs_demand_fee, 2),
                    "Oversize Demand Fee": round(oversize_demand_fee, 2),
                    "Fuel Fee": round(fuel_fee, 2),
                    "Repriced Total": round(total, 2),
                })

                result_rows.append(out)

                self.preview.insert(
                    "",
                    "end",
                    values=(
                        service,
                        zone,
                        actual_weight,
                        dim_weight,
                        billable_weight,
                        f"{base_rate:.2f}",
                        f"{ahs_fee:.2f}",
                        f"{das_fee:.2f}",
                        f"{oversize_fee:.2f}",
                        f"{residential_fee:.2f}",
                        f"{signature_fee:.2f}",
                        f"{general_demand_fee:.2f}",
                        f"{ahs_demand_fee:.2f}",
                        f"{oversize_demand_fee:.2f}",
                        f"{fuel_fee:.2f}",
                        f"{total:.2f}",
                    ),
                )

            out_df = pd.DataFrame(result_rows)

            final_cols = [
                "Transaction Date",
                "Invoice Date",
                "Invoice Number",
                "Tracking Number",
                "Service",
                "Zone",
                "Length",
                "Width",
                "Height",
                "Actual Weight",
                "DIM Weight",
                "Billable Weight",
                "Base Rate",
                "Additional Handling Surcharge",
                "DAS Fee",
                "Oversize Fee",
                "Residential Fee",
                "Signature Fee",
                "General Demand Fee",
                "AHS Demand Fee",
                "Oversize Demand Fee",
                "Fuel Fee",
                "Repriced Total",
            ]
            final_cols = [c for c in final_cols if c in out_df.columns]
            out_df = out_df[final_cols]

            numeric_cols = [
                "Zone", "Length", "Width", "Height", "Actual Weight", "DIM Weight", "Billable Weight",
                "Base Rate", "Additional Handling Surcharge", "DAS Fee", "Oversize Fee", "Residential Fee", "Signature Fee",
                "General Demand Fee", "AHS Demand Fee", "Oversize Demand Fee", "Fuel Fee", "Repriced Total",
            ]
            for col in numeric_cols:
                if col in out_df.columns:
                    out_df[col] = pd.to_numeric(out_df[col], errors="coerce").fillna(0)

            for col in ["Transaction Date", "Invoice Date"]:
                if col in out_df.columns:
                    out_df[col] = out_df[col].apply(
                        lambda v: pd.Timestamp(self._parse_date(v)) if self._parse_date(v) else pd.NaT
                    )

            if not output:
                raise UserInputError("Please choose an output file path.")

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                out_df.to_excel(writer, index=False, sheet_name="Repriced")

            self._apply_excel_formatting(output)

            self.output_path.set(output)
            self.status_var.set(f"Done. Saved to {output}")
            if hasattr(self, "run_summary_var"):
                self.run_summary_var.set(f"Processed {len(result_rows):,} shipment(s). Open the exported workbook for the complete breakdown.")
            messagebox.showinfo("Done", f"Repricing completed.\nSaved to:\n{output}")

        except UserInputError as e:
            self.status_var.set("Failed")
            messagebox.showerror("Error", str(e))
        except Exception as e:
            self.status_var.set("Failed")
            detail = traceback.format_exc(limit=5)
            messagebox.showerror("Unexpected Error", f"{e}\n\nTraceback:\n{detail}")


def main():
    root = ctk.CTk()
    try:
        ttk.Style(root).theme_use("clam")
    except Exception:
        pass
    FedExRepricingTool(root)
    root.mainloop()


if __name__ == "__main__":

    try:

        root = ctk.CTk()

        app = FedExRepricingTool(root)

        # =========================
        # AUTO LOAD CONFIG
        # =========================
        try:

            if os.path.exists(DEFAULT_CONFIG_PATH):

                app.config_path.set(DEFAULT_CONFIG_PATH)

                app.load_config()

        except Exception:
            pass

        # =========================
        # AUTO SAVE ON EXIT
        # =========================
        def on_closing():

            try:

                app.config_path.set(DEFAULT_CONFIG_PATH)

                app.save_config()

            except Exception:
                pass

            root.destroy()

        root.protocol("WM_DELETE_WINDOW", on_closing)

        root.mainloop()

    except Exception as e:

        error_text = traceback.format_exc()

        with open(
            "error_log.txt",
            "w",
            encoding="utf-8"
        ) as f:

            f.write(error_text)

        messagebox.showerror(
            "System Error",
            f"{e}\n\nDetailed error exported:\nerror_log.txt"
        )
